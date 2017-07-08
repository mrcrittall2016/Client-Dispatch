
# Python script to parse Excel macro Name.xlsm 
import openpyxl
import os
import re

# Data will be a list of rows represented as dictionaries
# containing each row's data.


#Function to populate dispatch sheet with compound data
def wednesday_dump(table_row, dispatch_sheet):
	
	#print "table_row is: %s" % table_row
	
	# Scan Dispatch sheet and insert
	for row in range(1, dispatch_sheet.max_row):
		
		# Get cells only for NBID column
		NBID_value = [dispatch_sheet.cell(row=row, column=col).value for col in range(1, 40) if col != None and col == 7]	
		
		print "NBID value is %r" % NBID_value
		
		
		#If cell is empty then store row
		if NBID_value == [None]:
									
			#print "Empty row number is: %d" % row
			
			# Now place from word doc into Excel 
			for key, value in table_row.iteritems():
				
				#print "They key is: %s" % key
				#print "The value is: %s" % value
				
				#print "row number is: %d" % row
				
				if key == 'Lab ID':
					dispatch_sheet.cell(row=row, column=7).value = value
				
				if key == 'Purity\n1dp':
					dispatch_sheet.cell(row=row, column=10).value = value
				
				if 'Rt' in key:
					dispatch_sheet.cell(row=row, column=11).value = value
					
				if 'Biosub' in key:
					dispatch_sheet.cell(row=row, column=2).value = value								
					
				if 'Store amount' in key:
					dispatch_sheet.cell(row=row, column=4).value = value
					
					if value == '':
						dispatch_sheet.cell(row=row, column=3).value = 'N'
						dispatch_sheet.cell(row=row, column=4).value = '-'
					
					else:
						dispatch_sheet.cell(row=row, column=3).value = 'Y'
					
				if 'DupOf' in key:
					dispatch_sheet.cell(row=row, column=19).value = value
					
				if 'Stereo' in key:
					dispatch_sheet.cell(row=row, column=21).value = value
					
				if 'ion' in key:
					dispatch_sheet.cell(row=row, column=13).value = value
					
				if 'Comments' in key:
					dispatch_sheet.cell(row=row, column=20).value = value
					
				if 'even' in key:
					dispatch_sheet.cell(row=row, column=29).value = value
				
				if 'odd' in key:
					dispatch_sheet.cell(row=row, column=30).value = value	
							
			
			break	
	
	


# Open individual dispatch sheet. Opening macro xlsm, so to avoid corrupting file need to send other arguments? : https://stackoverflow.com/questions/17675780/how-to-save-xlsm-file-with-macro-using-openpyxl
wb = openpyxl.load_workbook(filename='Name.xlsm', read_only=False, keep_vba=True)
sheet = wb.active 		

# Open Master Dispatch sheet
wb_dispatch = openpyxl.load_workbook('DISPATCH_TEMPLATE.xlsx')
dispatch_sheet = wb_dispatch.active 
					
print "Individual sheet is currently: %s" % sheet
print "Master dispatch sheet is: %s" % dispatch_sheet

# List of keys
keys = []

# Headers start from row 7. Seems that sheet.max_row represents last row with something in it
for row in range(7, sheet.max_row):	
	
	print "row is: %s" % row	
	
	# Count filled cells
	filled = 0		
	
	# List comprehension		
	single_row = [sheet.cell(row=row, column=col).value for col in range (1,16)]
	
	#print "\n" + str(single_row) + "\n"
	
	
	# Store headers as keys
	if row == 7:
		keys = single_row
		
	else:
		# Check how many empty cells are in a row
		for j in single_row:
			if j != None:
				filled += 1
		#print "\nFilled cells in this row are: %s\n" % filled
		
		# If less than 1 filled cell, discount row
		if filled < 1:
			single_row = []
		
	
		# Only zip row with keys list if they are the same length. Avoids mis-matched key-value pairs
		if len(single_row) == len(keys):				
			row_key_value_pairs = dict(zip(keys, single_row))				
			print "row %d is: %s " % (row, row_key_value_pairs)			
		
			# Now pass row to dispatch sheet
			wednesday_dump(row_key_value_pairs, dispatch_sheet)	
	
	
	if sheet.cell(row=row, column=col).value == None:
		continue
	
print "\n" + str(keys) + "\n"


# Gets all files of certain type in specified directory
#results = [each for each in os.listdir('/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Word_Excel') if each.endswith('.docx')]
wb.save('Name_test.xlsm')
wb_dispatch.save('Test.xlsx')
	
	
	
	
	
	
		
		