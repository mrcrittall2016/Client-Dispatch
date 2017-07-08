# http://stackoverflow.com/questions/27861732/parsing-of-table-from-docx-file

# Python script to parse Excel macro Name.xlsm 
import openpyxl
import os
import re
import datetime

# Data will be a list of rows represented as dictionaries
# containing each row's data.


#Function to populate dispatch sheet with compound data
def wednesday_dump(table_row, dispatch_sheet):
	
	#print "table_row is: %s" % table_row
	
	# Scan Dispatch sheet and insert
	for row in range(1, dispatch_sheet.max_row):
		
		# Get cells only for NBID column
		NBID_value = [dispatch_sheet.cell(row=row, column=col).value for col in range(1, 40) if col != None and col == 7]	
		
		print "NBID value is %r" % NBID_value
		
		
		#If cell is empty then store row
		if NBID_value == [None]:
									
			#print "Empty row number is: %d" % row
			
			# Now place from word doc into Excel 
			for key, value in table_row.iteritems():
				
				#print "They key is: %s" % key
				#print "The value is: %s" % value
				
				#print "row number is: %d" % row
				
				if key == 'Lab ID':
					dispatch_sheet.cell(row=row, column=7).value = value
				
				if key == 'Purity\n1dp':
					dispatch_sheet.cell(row=row, column=10).value = value
				
				if 'Rt' in key:
					dispatch_sheet.cell(row=row, column=11).value = value
					
				if 'Biosub' in key:
					dispatch_sheet.cell(row=row, column=2).value = value								
					
				if 'Store amount' in key:
					dispatch_sheet.cell(row=row, column=4).value = value
					
					if value == '':
						dispatch_sheet.cell(row=row, column=3).value = 'N'
						dispatch_sheet.cell(row=row, column=4).value = '-'
					
					else:
						dispatch_sheet.cell(row=row, column=3).value = 'Y'
					
				if 'DupOf' in key:
					dispatch_sheet.cell(row=row, column=19).value = value
					
				if 'Stereo' in key:
					dispatch_sheet.cell(row=row, column=21).value = value
					
				if 'ion' in key:
					dispatch_sheet.cell(row=row, column=13).value = value
					
				if 'Comments' in key:
					dispatch_sheet.cell(row=row, column=20).value = value
					
				if 'even' in key:
					dispatch_sheet.cell(row=row, column=29).value = value
				
				if 'odd' in key:
					dispatch_sheet.cell(row=row, column=30).value = value	
							
			
			break	

def create_dictionary(sheet, dispatch_sheet, date, day):
	# Headers start from row 7. Seems that sheet.max_row represents last row with something in it
	for row in range(7, sheet.max_row):	

		print "row is: %s" % row	

		# Count filled cells
		filled = 0		

		# List comprehension		
		single_row = [sheet.cell(row=row, column=col).value for col in range (1,16)]

		#print "\n" + str(single_row) + "\n"


		# Store headers as keys
		if row == 7:
			keys = single_row

		else:
			# Check how many empty cells are in a row
			for j in single_row:
				if j != None:
					filled += 1
			#print "\nFilled cells in this row are: %s\n" % filled

			# If less than 1 filled cell, discount row
			if filled < 1:
				single_row = []


			# Only zip row with keys list if they are the same length. Avoids mis-matched key-value pairs
			if len(single_row) == len(keys):				
				row_key_value_pairs = dict(zip(keys, single_row))				
				print "row %d is: %s " % (row, row_key_value_pairs)			
				
				if day == "Wednesday":
					# Now pass row to dispatch sheet
					wednesday_dump(row_key_value_pairs, dispatch_sheet)			


		if sheet.cell(row=row, column=col).value == None:
			continue

		print "\n" + str(keys) + "\n"



# Function to grab a row at a time from each person's dispatch sheet
def analyse_dispatch_sheets(Macro_docs, dispatch_sheet, date, Day, day_number):
	
	
	for doc in Macro_docs:		
		
		print "doc is: %s" % doc
		
		# Open individual dispatch sheet. Opening macro xlsm, so to avoid corrupting file need to send other arguments? : https://stackoverflow.com/questions/17675780/how-to-save-xlsm-file-with-macro-using-openpyxl
		wb = openpyxl.load_workbook(filename=doc, read_only=False, keep_vba=True)
		sheet = wb.active 
					
		print "Individual sheet is currently: %s" % sheet
		print "Master dispatch sheet is: %s" % dispatch_sheet
		
		create_dictionary(sheet, dispatch_sheet, date, Day)

		wb.save(doc)
		


# Empty array for storing dictionary
keys =[]	
		
# Get today's date
print('Date today: %s' % datetime.date.today())
date = datetime.date.today()

# Determines and prints actual day of the week and month. See here for documentation: https://docs.python.org/3/library/datetime.html#strftime-strptime-behavior
now = datetime.datetime.now()

# Get the day of the week
Day = now.strftime("%A")
#print "The day of the week is: %s" % Day

# Get the month of the year
Month = now.strftime("%B")
#print"The month is: %s" % Month

# Get the current year
Year = now.strftime("%Y")
#print"The year is: %s" % Year

# What is the day of the month
day_number = now.strftime("%d")
#print "The day of the month is: %s" % day_number

# Gets all word doc files in specified directory
Macro_docs = [each for each in os.listdir("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch") if each.endswith('.xlsm')]

# If no dispatch sheets available, quit program
if not Macro_docs:
	print "No Dispatch sheets available for analysis"	
	quit()

print "Current dispatch sheets are: %s" % Macro_docs
	
#Day = "Thursday"
Day = raw_input("Please provide the day of the week: ")

if Day.isalpha() == False:
	print "Please provide a valid day of the week"


# Wednesday
if Day == 'Wednesday':
	
	#print "It is Wednesday"	
	
	# Open Master Dispatch sheet
	wb_dispatch = openpyxl.load_workbook('DISPATCH_TEMPLATE.xlsx')
	dispatch_sheet = wb_dispatch.active 
	
	# If Wednesday then need to create new dispatch excel sheet based on timedate stamp
	#print "Saving new spreadsheet in appropiate folder... "
	
	# First create appropiate directory if does not exist...
	if not os.path.exists("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch/%s/%s_%s" % (Year, Month, Year)):
		os.makedirs("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch/%s/%s_%s" % (Year, Month, Year))	
	
			
	# Add 1 to day number (dispatch day is next day ie Wednesday)
	#day_number = int(day_number) + 1
	#day_number = "%02d" % day_number
	
	analyse_dispatch_sheets(Macro_docs, dispatch_sheet, date, Day, day_number)
	
	#print "day_number is: %s" % day_number	
	
	# Note that day number needs to be +1
	#save_location = "s:/Receptos/Dispatch/%s/%s_%s/%s_%s_%s.xlsx" % (Year, Month, Year, day_number, Month, Year) 
	save_location = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch/%s/%s_%s/%s_%s_%s.xlsx" % (Year, Month, Year, day_number, Month, Year)
		
	wb_dispatch.save(save_location)
	
	print"\nProcess completed successfully\n"
	
	#Finish
	quit()








	
	
	
	
	
		
		