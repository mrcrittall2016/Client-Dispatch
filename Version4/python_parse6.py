# http://stackoverflow.com/questions/27861732/parsing-of-table-from-docx-file
# 12-07-2017: Added to Wednesday dump inputs for Project, company, no_salt
# 12-07-2017: Trying to connect Python with VBA... try passing array instead of looping through array in python
# 16-07-2017: Populate dispatch sheet template with individual_dispatch ids instead 
# 19-07-2017: Now after developing new individual dispatch sheet to enable copying of structures across on Wednesday, this version will re-introduce Thursday DMPK as well as extracting information for FedX for Wednesday

# Python script to parse Excel macro Name.xlsm 
import openpyxl
import sys
#import win32com.client
import os
import re
import logging
import datetime
from datetime import time
from shutil import copyfile 
import itertools #This module is required for iterating through two arrays/objects in parallel. See "copy_across_sheet_ids"

# Data will be a list of rows represented as dictionaries.
# Function to drive VBA macro from Python for copying and pasting chem structures
def thursday_VBA(source, target):

	Excel_macro_path = "control_files/DMPK_control.xlsm"
	Excel_macro = "DMPK_control.xlsm"	
	files = [source, target, Excel_macro_path]
	
	#Loop through file locations to check can find everything. If not, exit program
	for file in range(1, len(files)):
		return_check = os.path.exists(files[file])		
		if return_check == False:
			print "\nSorry, unable to access %s\n" % files[file]
			quit()
	
	print "\nAll files accounted for...\n"
	
	try:			
		#return_check = os.path.exists('s:\Receptos\Dispatch\python\control_files\%s' % Excel_macro)
		#print "The return value of os_check for VBA macro is: %s" % return_check			
		xl = win32com.client.Dispatch("Excel.Application")
		#print "Application is: %s" % xl
		xl.Workbooks.Open(Filename="%s" % Excel_macro_path, ReadOnly=1)
	except Exception, e:
		print "\nSorry, unable to open Excel macrobook: \n" + str(e)
	
	try:
		#Run macro. Hard-code source and target files first to test then, pass as variables. Works hard-coded 23/06/2017 
		xl.Application.Run("%s!DMPK.import_structures_to_DMPK" % Excel_macro, source, target)
		print "\nMacro was run successfully!\n"
	except Exception, e:
		print "\nSorry, unable to run macro\n" + str(e)		
	
		#Run macro. Hard-code source and target files first to test then, pass as variables 
		#xl.Application.Run("Dispatch_control-V3.xlsm!ModDataExport.import_structures_to_DMPK", dispatch_sheet_source, DMPK_sheet_target)
		#xl.Application.Run("Dispatch_control-V3.xlsm!ModDataExport.import_structures_to_DMPK")
	finally:		
		xl.Application.Quit()
		del xl




# Function to return number of barcodes and hence vials as well as largest quantity and barcode range. See here on how to return multiple values: http://stackoverflow.com/questions/354883/how-do-you-return-multiple-values-in-python
# Return an object (or class) 
class FedX_data(object):
	
	def __init__(self,	number_of_vials, biosub_barcode_range, store_barcode_range, highest_quantity):
		self.number_of_vials = number_of_vials
		self.biosub_barcodes = biosub_barcode_range
		self.store_barcodes = store_barcode_range
		self.highest_quantity = highest_quantity
	
	def summarise_data(self):
		
		return "The total number of vials for dispatch is: %s \n\nThe biosub barcode range is: %s \n\nThe store barcode range is: %s \n\nThe highest quantity of material for dispatch is: %s mg\n\n" % (self.number_of_vials, self.biosub_barcodes, self.store_barcodes, self.highest_quantity)
		
	
		
# Go through dispatch sheet and calculate number of vials, barcode range and highest quantity of material for dispatch
def FedX(sheet):
	
	# List for storing barcodes and masses
	biosub_barcodes = []
	store_barcodes = []
	masses = []
	
	# Ignore header
	for row in range(2, sheet.max_row):		
			
		# Get cell_value only for barcode columns
		barcode_even = ''.join([sheet.cell(row=row, column=col).value for col in range(1, 40) if sheet.cell(row=row, column=col).value != None and col == 29])
		
		barcode_odd = ''.join([sheet.cell(row=row, column=col).value for col in range(1, 40) if sheet.cell(row=row, column=col).value != None and col == 30])
				
		# Note: If retrieving floats from Excel sheet, have to convert each member of list to string individually. Do this using map function: https://stackoverflow.com/questions/6507431/join-float-list-into-space-separated-string-in-python
		sample_test = ''.join(map(str,[sheet.cell(row=row, column=col).value for col in range(1, 40) if sheet.cell(row=row, column=col).value != None and col == 2]))		
		
		sample_store = ''.join(map(str,[sheet.cell(row=row, column=col).value for col in range(1, 40) if sheet.cell(row=row, column=col).value != None and col == 4]))
				
	
		if barcode_even != '':
			biosub_barcodes.append(barcode_even)
			
		if barcode_odd != '':			
			store_barcodes.append(barcode_odd)
		
		if sample_test != '':			
			masses.append(float(sample_test))			
		
		if sample_store != '' and sample_store != '-':			
			masses.append(float(sample_store))
			
			
		
		# If hit row where encounter two empty barcodes then this must be the end of the sheet. Confirm by checking NBID cell is also empty  
		if barcode_even == '' and barcode_odd == '' and sheet.cell(row=row, column=7).value == None:
			break
	
			
	# Sort barcode lists into ascending order and check
	biosub_barcodes.sort()
	store_barcodes.sort()
	#print barcodes	
	
	# Calculate number of vials
	Number_of_vials = len(biosub_barcodes) + len(store_barcodes)
	
	#print "Number of vials being sent is: %d" % Number_of_vials
	
	biosub_barcode_range = "%s to %s" % (biosub_barcodes[0], biosub_barcodes[len(biosub_barcodes) - 1])
	store_barcode_range = "%s to %s" % (store_barcodes[0], store_barcodes[len(store_barcodes) - 1])
	#print barcode_range
	
	#Check masses array
	#print "Mass array is: %s" % masses
	
	#Obtain max quantity
	max_quantity = max(masses)
	#print "Max quant. is %s: " % max_quantity
	
	#print "Maximum quantity is: %s mg" % max_quantity
	
	# Now return class
	return_class = FedX_data(Number_of_vials, biosub_barcode_range, store_barcode_range, max_quantity)
	return return_class


#Function to place sheet ids in dispatch template
def copy_across_sheet_ids(dispatch_template, individual_sheets):
	
	# Required to iterate over rows of dispatch sheet and individual sheet array simultaneously. See here: https://stackoverflow.com/questions/1663807/how-can-i-iterate-through-two-lists-in-parallel
	for (f,b) in itertools.izip_longest(range(1, len(individual_sheets)+1), individual_sheets):
		#print "f: ", f ,"; b: ", b
		if f != None:		
			dispatch_template.cell(row=f+1, column=1).value = b
		

#Function to populate dispatch sheet with compound data
def wednesday_dump(table_row, dispatch_sheet):
	
	#print "table_row is: %s" % table_row
	
	# Scan Dispatch sheet and insert
	for row in range(1, dispatch_sheet.max_row):	
		
		
		# Add Project for column 5
		dispatch_sheet.cell(row=row, column=5).value = 'Project'
		
		# Add Company for column 6
		dispatch_sheet.cell(row=row, column=6).value = 'Company'
		
		# Add XSel15min for column 12
		dispatch_sheet.cell(row=row, column=12).value = 'XSel15min'
		
		# Add 'no_salt' for column 14
		dispatch_sheet.cell(row=row, column=14).value = 'no_salt'		
		
		# Get cells only for NBID column
		NBID_value = [dispatch_sheet.cell(row=row, column=col).value for col in range(1, 40) if col != None and col == 7]	
		
		#print "NBID value is %r" % NBID_value
		
		
		#If cell is empty then store row
		if NBID_value == [None]:
									
			#print "Empty row number is: %d" % row
			
			# Now place from word doc into Excel 
			for key, value in table_row.iteritems():
				
				#print "They key is: %s" % key
				#print "The value is: %s" % value
				
				#print "row number is: %d" % row
				
				if key == 'Lab ID':
					dispatch_sheet.cell(row=row, column=7).value = value
				
				if 'Purity' in key and '1 dp' in key:
					dispatch_sheet.cell(row=row, column=10).value = value
				
				if 'Rt' in key:
					dispatch_sheet.cell(row=row, column=11).value = value
					
				if 'Biosub' in key:
					dispatch_sheet.cell(row=row, column=2).value = value								
					
				if 'Store amount' in key:
					dispatch_sheet.cell(row=row, column=4).value = value
					
					if value == '':
						dispatch_sheet.cell(row=row, column=3).value = 'N'
						dispatch_sheet.cell(row=row, column=4).value = '-'
					
					else:
						dispatch_sheet.cell(row=row, column=3).value = 'Y'
					
				if 'DupOf' in key:
					dispatch_sheet.cell(row=row, column=19).value = value
					
				if 'Stereo' in key:
					dispatch_sheet.cell(row=row, column=21).value = value
					
				if 'Ion' in key:
					dispatch_sheet.cell(row=row, column=13).value = value
					
				if 'Comments' in key:
					dispatch_sheet.cell(row=row, column=20).value = value
					
				if 'Syg' in key:
					dispatch_sheet.cell(row=row, column=29).value = value
				
				if 'Celgene' in key:
					dispatch_sheet.cell(row=row, column=30).value = value
							
			
			break	

# Function to scan DMPK spreadsheet and fill in appropiate Excel cells
def thursday_DMPK(dictionary, DMPK_sheet, date):	
	
	# Get value of DMPK Box Grid from dictionary and convert to string 
	reference = ''.join([dictionary[key] for key in dictionary if 'Box' in key])		
	#print "reference from individual sheet is: %s" % reference
	
	# Note, may need to keep an eyes on this column reference CHANGE BACK
	grid_reference = DMPK_sheet.cell(row=row, column=6).value
	#print "reference in DMPK_template is: %s" % grid_reference
	
	# If can't get or find grid_referece.. return
	if grid_reference == None:
		return False 
	
	# Remove first 5 characters from string ie 'Box#/' See: http://stackoverflow.com/questions/11806559/removing-first-x-characters-from-string	
	grid_reference = grid_reference[5:]
	#print "excel reference is: %s" % grid_reference	
	
	# Get amount of sample submitted. Note, when dealing with a list of floats from Excel, need to use map function to convert each entry of the list to a string, not the whole list at once. See here: https://stackoverflow.com/questions/6507431/join-float-list-into-space-separated-string-in-python
	sample_quantity = ''.join(map(str, [dictionary[key] for key in dictionary if 'DMPK' in key and '2 dp' in key]))
			
	# Get ELN and convert to string
	ELN = ''.join([dictionary[key] for key in dictionary if 'ID' in key])
	
	#print "word doc reference is: %s" % reference
	
	# Bug here - if people abbreviate Box to ie B2 or B3 then breaks when look for Excel grid ref of B2 or B3 in word doc. Also need to figure out way to exit function if insert compound instead of still scanning through.. return true and false?
	# To fix bug - look for '/' and if finds, remove any text before and set as new reference. OR, test to see if first two characters are B2 or B3. If so, ignore.
	
	# Check first 2 characters of reference if equal grid_reference, delete
	if grid_reference == reference[:2]:
		reference = reference[2:]	
	
		
	# If the Excel grid ref is found in the word doc grid ref 
	if grid_reference in reference:
		print "Found it!"		
		
		# Then copy across appropiate information
		if DMPK_sheet.cell(row=row, column=5).value == None:
			DMPK_sheet.cell(row=row, column=5).value = ELN
			
		if DMPK_sheet.cell(row=row, column=2).value == None:
			DMPK_sheet.cell(row=row, column=2).value = sample_quantity
			
		if DMPK_sheet.cell(row=row, column=7).value == None:
			DMPK_sheet.cell(row=row, column=7).value = date
			
		return True
	
	return False			
	
		
def create_dictionary(individual_sheet, dispatch_sheet_or_DMPK_sheet, date, day):
	# Headers start from row 7. Seems that sheet.max_row represents last row with something in it
	
	#print "dispatch_sheet/DMPK sheet is: %s" % dispatch_sheet_or_DMPK_sheet
	#print "individual sheet is: %s" % individual_sheet
	
	for row in range(7, individual_sheet.max_row):	

		#print "row is: %s" % row	

		# Count filled cells
		filled = 0		

		# List comprehension		
		single_row = [individual_sheet.cell(row=row, column=col).value for col in range (1,16)]

		#print "\n" + str(single_row) + "\n"


		# Store headers as keys
		if row == 7:
			keys = single_row

		else:
			# Check how many empty cells are in a row
			for j in single_row:
				if j != None:
					filled += 1
			#print "\nFilled cells in this row are: %s\n" % filled

			# If less than 1 filled cell, discount row
			if filled < 1:
				single_row = []


			# Only zip row with keys list if they are the same length. Avoids mis-matched key-value pairs
			if len(single_row) == len(keys):				
				row_key_value_pairs = dict(zip(keys, single_row))				
				#print "row %d is: %s " % (row, row_key_value_pairs)			
				
				if day == "Wednesday":
					# Now pass row to dispatch sheet
					wednesday_dump(row_key_value_pairs, dispatch_sheet_or_DMPK_sheet)			
			
			
				# Only zip row with keys list if they are the same length. Avoids mis-matched key-value pairs			
				if day == "Thursday":					
					global compounds					
					# Need to only increment compounds if DMPK Box/grid column is filled					
					check = DMPK_row_check(row_key_value_pairs)					
							
						
					if check == True:					
						compounds += 1															
					
					return_value = thursday_DMPK(row_key_value_pairs, dispatch_sheet_or_DMPK_sheet, date)			
					
					
					# Only increment inserted if return_value is True ie finds reference in word doc
					if return_value == True:
						global inserted
						inserted +=1
						# Exit function if have found compound. No point in looking anymore
						#return 'found' 
					
			

		if individual_sheet.cell(row=row, column=col).value == None:
			continue

		#print "\n" + str(keys) + "\n"



# Function to grab a row at a time from each person's dispatch sheet
def analyse_dispatch_sheets(Macro_docs, dispatch_sheet_or_DMPK_sheet, date, Day, day_number):
	
	# Variables to keep track of how many and which files are still open
	open_file_count = 0
	open_docs = []	
	
	# This bit of code (move to separate function) analyses the word_docs list for any copies... ie files created if the orginal file is open. As found that even if move the temp file, still can not manipulate the file that is open. So have to create a copy and manipulate the copy instead.
	# Manipulating the copy does work! (Test if still need to delete the temp as well as create copy - No, do not have to move the temp. Can just exclude from the word_list array)
	
	
	# Use copy of files that are already open. Delete after finished with them
	for doc in Macro_docs:		
				
		if 'copy' in doc:			
			# Get name of file before copy
			name = doc.rsplit('_copy', 1)[0]			
			# Now ignore/remove that file from word_docs list
			name = name + '.xlsm'
			#print name			
			Macro_docs.remove(name)
	
	
	for doc in Macro_docs:		
		
		# Open individual dispatch sheet. Opening macro xlsm, so to avoid corrupting file need to send other arguments? : https://stackoverflow.com/questions/17675780/how-to-save-xlsm-file-with-macro-using-openpyxl
		wb = openpyxl.load_workbook(filename=doc, read_only=False, keep_vba=True)
		individual_sheet = wb.active 			
		
		
		create_dictionary(individual_sheet, dispatch_sheet_or_DMPK_sheet, date, Day)
		
		
		# save document 
		try:
			wb.save(doc)
			# If doc contains 'copy', delete it
			if 'copy' in doc:
				os.remove("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % doc)
		
		except:
			
			print "\nSorry, the file %s is still open..." % doc
			
			# Update local open_file_count
			open_file_count += 1
			# Push to open_doc array
			open_docs.append(doc)	
		
		# If more than one document is open, offer user choices as to copy files or not. Then quit the program
		if open_file_count > 0:
				
			naughty_list = ' and '.join(open_docs)
		
			if now_time >= time (10,45):
		
				print"\nPlease check with %s to see if they are finished with their word document(s)...\n" % naughty_list
			
				return_value = raw_input("Please type 'yes' if they are finished, 'no' if the individual(s) requires more time: ")
			
				if return_value == 'yes':
					check = raw_input("Would you like to create a copy of the file(s)?: ")					
					if check == 'yes':
						# Try to copy and rename open files (and then delete copy afte finished with it?) and run again. Would be great here to be able to automatically run script again
						for file in open_docs:
							copy = file.replace('.xlsm','') + '_copy.xlsm'								
							copyfile("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % file, "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % copy)
							print "\nCopy %s has been made.\n" % copy	
						print "\nPlease re-run the script\n"					
					else:
						print "\nOk, please try and run the script again in a few moments\n"
			
				else:
					print "\nOk, please try again later\n"
		
			else:
				print "\nPlease allow more time for completion of dispatch sheets\n"		
		
			quit()	

# Checks to see if cell value of DMPK columns are filled or not ie if contains '-' or ''
def DMPK_row_check(row_key_value_pairs):	
	for key, value in row_key_value_pairs.iteritems():
		if 'Box' in key:			
			
			if value == None:
				print "Value is None"
				return False
				
			if '-' in value:
				return False
			
			
	return True
		

# Variables for file/directory locations so as to avoid hard-coding
individual_dispatch_sheets_location = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_git/Version4"
master_dispatch_sheet_template = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_git/Version4/control_files/DISPATCH_TEMPLATE_new.xlsm"
where_will_master_dispatch_be_saved = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_git/Version4"
DMPK_folder = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_git/Version4/DMPK_files"
new_DMPK_sheet_template = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_git/Version4/control_files/DMPKsamples_TEMPLATE.xlsx"

# Empty array for storing dictionary
keys =[]

# Global variable for counting number of coupounds
compounds = 0
inserted = 0	
		
# Get today's date
print('Date today: %s' % datetime.date.today())
date = datetime.date.today()

# Determines and prints actual day of the week and month. See here for documentation: https://docs.python.org/3/library/datetime.html#strftime-strptime-behavior
now = datetime.datetime.now()
now_time = now.time()

# Get the day of the week
Day = now.strftime("%A")
#print "The day of the week is: %s" % Day

# Get the month of the year
Month = now.strftime("%B")
#print"The month is: %s" % Month

# Get the month_number ie March should be 3
month_number = now.strftime("%m")

# Strip off front zero 
month_number = month_number.lstrip("0")

# Get the current year
Year = now.strftime("%Y")
#print"The year is: %s" % Year

# What is the day of the month
day_number = now.strftime("%d")
#print "The day of the month is: %s" % day_number

# Gets all Excel macro files in specified directory
Macro_docs = [each for each in os.listdir(individual_dispatch_sheets_location) if each.endswith('.xlsm')]

# Exclude temp files
Macro_docs = [doc for doc in Macro_docs if not doc.startswith('~')]	

# If no dispatch sheets available, quit program
if not Macro_docs:             
	print "\nNo Dispatch sheets available for analysis\n"	
	quit()

print "\nCurrent dispatch sheets are: %s\n" % Macro_docs
	
#Day = "Thursday"
Day = raw_input("\nPlease provide the day of the week: ")

if Day.isalpha() == False:
	print "\nPlease provide a valid day of the week\n"


# Wednesday
if Day == 'Wednesday':
	
	#print "It is Wednesday"	
	
	# Open Master Dispatch sheet
	wb_dispatch = openpyxl.load_workbook(filename=master_dispatch_sheet_template, read_only=False, keep_vba=True)
	dispatch_sheet = wb_dispatch["ChemOffice1"]
	
	# If Wednesday then need to create new dispatch excel sheet based on timedate stamp
	
	# First create appropiate directory if does not exist...
	if not os.path.exists("%s/%s/%s_%s_%s" % (where_will_master_dispatch_be_saved, Year, month_number, Month, Year)):
		os.makedirs("%s/%s/%s_%s_%s" % (where_will_master_dispatch_be_saved, Year, month_number, Month, Year))	
		
	analyse_dispatch_sheets(Macro_docs, dispatch_sheet, date, Day, day_number)	
	
	#Add source_sheet names to separate sheet of dispatch template
	copy_sheet_ids = wb_dispatch["Import_Structures"]	
	copy_across_sheet_ids(copy_sheet_ids, Macro_docs)	 	
	
	
	# Only generate FedX info if after 10:45. See here: http://stackoverflow.com/questions/10048249/how-do-i-determine-if-current-time-is-within-a-specified-range-using-pythons-da
	if now_time >= time(4,45) and now_time <= time(16,00):
		# Need a function to calculate barcode range, number of compounds, number of vials and highest quantity being sent
		Dispatch_data = FedX (dispatch_sheet)
		summary = Dispatch_data.summarise_data()		
			
		#summary_file = open("S:/Receptos/Team_individual_folders/Matt_C/VBA_test/new_sheets/%s/%s_%s_%s/%s %s %s/FedX_info.txt" % (Year, month_number, Month, Year, day_number, Month, Year), "w")
		summary_file = open("%s/%s/%s_%s_%s/FedX_info.txt" % (where_will_master_dispatch_be_saved, Year, month_number, Month, Year), "w")					
		summary_file.write(summary)	
		summary_file.close()	
	

	
	#save_location = "s:/Receptos/Dispatch/%s/%s_%s/%s_%s_%s.xlsx" % (Year, Month, Year, day_number, Month, Year) 
	save_location = "%s/%s/%s_%s_%s/%s_%s_%s.xlsm" % (where_will_master_dispatch_be_saved, Year, month_number, Month, Year, day_number, Month, Year)		
	wb_dispatch.save(save_location)
	#copy_in_structures(Macro_docs, save_location, Day) #Perhaps only run this function at end of day i.e. before uplaoding spreadsheet. Seems like could be a bit buggy... so don't want to screw up in morning ie keep seeing 'runtime error 13 type mismatch'. Perhaps because old file has not closed? Trying to open ~file?
	#Perhaps opening and closing comm link for every Doc is also causing problems. Perhaps better to pass Excel array of files and let VBA iterate through the array rather than Python
		 	
	
	print"\nProcess completed successfully\n"
	
	#Finish
	quit()


# THURSDAY
if Day == 'Thursday':
	
	# If day is Thursday then need to retrieve latest DMPK Box and ask user for date of last Dispatch
	last_wednesday_dispatch_date = raw_input("\nPlease provide the date of Wednesday's Dispatch in the following format: day_month_year. For example, 05_July_2017: ")
	
	#Later, will have to extract Month, day_number and Year from last_wednesday_dispatch_date variable
	
	latest_dispatch_sheet = "%s/%s/%s_%s_%s/%s.xlsm" % (where_will_master_dispatch_be_saved, Year, month_number, Month, Year, last_wednesday_dispatch_date) 
	
	print latest_dispatch_sheet
	
	# Test to see if can reach file... if can not, quit and ask user to run program again
	ret = os.access(latest_dispatch_sheet, os.F_OK)	 
	
	
	if ret == True:		
		print "\nDispatch sheet found, copying across data to latest DMPK sheet\n"		
		#Convert dispatch date to that to be presented in spreadsheet
		dispatch_date = last_wednesday_dispatch_date.replace("_", "/")	
		print dispatch_date		
		
		DMPK_boxes = [each for each in os.listdir(DMPK_folder) if each.endswith('.xlsx')]
				
		# Get int from latest DMPK box just in case have to create a new sheet... see: http://stackoverflow.com/questions/10365225/extract-digits-in-a-simple-way-from-a-python-string
		box_number = ''.join(re.findall('\d+', max(DMPK_boxes)))
				
		# Load latest DMPK box
		latest_DMPK = "%s/%s" % (DMPK_folder, max(DMPK_boxes))
		print "Latest dispatch DMPK sheet is: %s" % latest_DMPK	
		
		wb_DMPK = openpyxl.load_workbook(latest_DMPK)		
		DMPK_sheet = wb_DMPK.active		
		
		# Find first blank cell in ELN column... will need to remove + 1 later CHANGE BACK
		for row in range(1, DMPK_sheet.max_row + 1):
			
						
			# Get cells only for NBID column
			ELN_ref = [DMPK_sheet.cell(row=row, column=col).value for col in range(1, 10) if col != None and col == 5]	
						
			# Find first blank cell
			if ELN_ref == [None]:
				
				print "compounds inserted so far: %s" % inserted
				#print sheet.cell(row=row, column=6).value
				
				# Ensure re-set global compound count before each analyse is called is called			
				compounds = 0
				
				# Call analyse_word_docs here to zip up each row in word_docs and find grid reference amoungst all sheets. Slower alogrithm, but neccessary
				analyse_dispatch_sheets(Macro_docs, DMPK_sheet, dispatch_date, Day, day_number)

				# Once inserted all compounds, save and close workbook
				print "compound total is: %s" % compounds
				
				# If have inserted all compounds, finish loop and copy structures across. 
				if inserted == compounds:	
					
					save_DMPK = "%s/%s" % (DMPK_folder, max(DMPK_boxes))				
					wb_DMPK.save(save_DMPK)					
					#thursday_VBA(latest_dispatch_sheet, save_DMPK)				
					break				
				
				#print "Checking grid ref: %s" % DMPK_sheet.cell(row=row, column=6).value
											
				
				# If hit H9.. that is end of sheet
				if 'H9' in DMPK_sheet.cell(row=row, column=6).value:
					print 'found last cell'				
					
					#If inserted is 0 then sheet must actually already be full
				 	if inserted == 0:
						print "\nSheet is already complete for this week!\n"
						quit()
					
					
					# Are there still more compounds to add? If so, must need to creat a new box. So need to know number of compounds... this should be equivalent to total number of rows from word docs.	
					# If number of compounds inserted is less than total number to be inserted...
					if inserted < compounds:
						
						print "Need to create another spreadsheet..."
						
						# First close current worksheet...and copy structures
						save_DMPK = "%s/%s" % (DMPK_folder, max(DMPK_boxes))						
						wb_DMPK.save(save_DMPK)						
						#thursday_VBA(latest_dispatch_sheet, save_DMPK)
						
						# Add 1 to latest box_number...then convert back to string
						box_new = int(box_number) + 1
						box_new = str(box_new)
						
						# Now create new spreadsheet
						filepath = new_DMPK_sheet_template 
						wb = openpyxl.load_workbook(filepath)
						new_sheet_DMPK = wb.active				
						
						# Populate column 6 ie Box well# by find and replace. For new sheet, want to go through whole thing either filling with new compounds or replace hash with number. Have this as separate function? ie fill_new_sheet()
						# CHANGE and remove -1
						for row in range(1, new_sheet_DMPK.max_row + 1):
							# Get cells only for Box/grid column
							Box_grid = [new_sheet_DMPK.cell(row=row, column=col).value for col in range(1, 10) if col != None and col == 6]
							
							ELN_ref = [new_sheet_DMPK.cell(row=row, column=col).value for col in range(1, 10) if col != None and col == 5]
							
							# Convert list to string
							Box_grid = ''.join(Box_grid)						
							
							# Replace # with box_new
							Box_grid = Box_grid.replace('#', box_new)
							new_sheet_DMPK.cell(row=row, column=6).value = Box_grid		

							#print Box_grid
							#print ELN_ref
							
							if ELN_ref == [None]:
																						
								# Only insert a compound if total compound count has not been reached
								if inserted != compounds:
									
									#inserted += 1
									compounds = 0
									# Call analyse_word_docs here to zip up word_docs as dictionary rows and find grid reference amoungst all sheets. Slower alogrithm, but neccessary
									analyse_dispatch_sheets(Macro_docs, new_sheet_DMPK, dispatch_date, Day, day_number)
									
									# Uncode this to monitor what is going on
									print "In new sheet, total compounds inserted is: %s" % inserted
									print "In new sheet, compounds is: %s" % compounds
							
							
							# If reach end of worksheet
							if 'H9' in Box_grid:
								print "H9 here"							
								# Save new worksheet and copy across structures
								new_save = "%s/DMPKsamples_BOX%s.xlsx" % (DMPK_folder, box_new)							
								wb.save(new_save)								
								#thursday_VBA(latest_dispatch_sheet, new_save)
								break					
						
						
					break 
				
			
		# Now clean up word documents
		#clean_up(Macro_docs, Year, day_number, Month)	
		
		#Finish
		print "\nProcess completed successfully\n"
		quit()
		
		
	else:
		print "\nSorry, dispatch sheet could not be found, please try re-entering date\n"
		quit()

else:
	print "\nScript is not due to run today. Please input either Wednesday or Thursday to run\n"






	
	
	
	
	
		
		