# http://stackoverflow.com/questions/27861732/parsing-of-table-from-docx-file
# 12-07-2017: Added to Wednesday dump inputs for CXCR3, company, no_salt
# 12-07-2017: Trying to connect Python with VBA... try passing array instead of looping through array in python
# 16-07-2017: Populate dispatch sheet template with individual_dispatch ids instead 

# Python script to parse Excel macro Name.xlsm 
import openpyxl
import sys
#import win32com.client
import os
import re
import logging
import datetime
from datetime import time
from shutil import copyfile 
import itertools

# Data will be a list of rows represented as dictionaries.


#Function to copy structures using VBA. win32com.client does not work on Mac unfortunately
def copy_in_structures(source, target, Day):
	
	print "Current source is: %s" % source
	
	if Day == "Wednesday":		
		#for source_doc in source:	
		Excel_macro_path = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/Dispatch_control.xlsm"
		Excel_macro = "Dispatch_control.xlsm"	
		
		#Provide absolute path for Excel dispatch sheets
		#source = "s:/Receptos/Team_individual_folders/Matt_C/VBA_test/new_sheets/" + source
					
		files = [target, Excel_macro_path]		
		
		#Loop through file locations to check can find everything. If not, exit program
		for file in range(1, len(files)):				
			return_check = os.path.exists(files[file])		
			if return_check == False:
				print "\nSorry, unable to access %s\n" % files[file]
				quit()
		
		print "\nAll files accounted for...\n"
		
		try:			
			#return_check = os.path.exists('s:\Receptos\Dispatch\python\control_files\%s' % Excel_macro)
			#print "The return value of os_check for VBA macro is: %s" % return_check			
			xl = win32com.client.Dispatch("Excel.Application")
			#print "Application is: %s" % xl
			xl.Workbooks.Open(Filename="%s" % Excel_macro_path, ReadOnly=1)
		except Exception, e:
			print "\nSorry, unable to open Excel macrobook: \n" + str(e)
		
		try:
			#Run macro. Hard-code source and target files first to test then, pass as variables. Works hard-coded 23/06/2017 
			xl.Application.Run("%s!Dispatch2.import_structures_to_Dispatch" % Excel_macro, source, target)
			print "\nMacro was run successfully!\n"
		except Exception, e:
			print "\nSorry, unable to run macro\n" + str(e)		
		
			#Run macro. Hard-code source and target files first to test then, pass as variables 
			#xl.Application.Run("Dispatch_control-V3.xlsm!ModDataExport.import_structures_to_DMPK", dispatch_sheet_source, DMPK_sheet_target)
			#xl.Application.Run("Dispatch_control-V3.xlsm!ModDataExport.import_structures_to_DMPK")
		finally:		
			xl.Application.Quit()
			del xl


#Function to place sheet ids in dispatch template
def copy_across_sheet_ids(dispatch_template, individual_sheets):
	
	# Required to iterate over rows of dispatch sheet and individual sheet array simultaneously. See here: https://stackoverflow.com/questions/1663807/how-can-i-iterate-through-two-lists-in-parallel
	for (f,b) in itertools.izip_longest(range(1, len(individual_sheets)+1), individual_sheets):
		print "f: ", f ,"; b: ", b
		if f != None:		
			dispatch_template.cell(row=f, column=1).value = b
		

#Function to populate dispatch sheet with compound data
def wednesday_dump(table_row, dispatch_sheet):
	
	#print "table_row is: %s" % table_row
	
	# Scan Dispatch sheet and insert
	for row in range(1, dispatch_sheet.max_row):	
		
		
		# Add CXCR3 for column 5
		dispatch_sheet.cell(row=row, column=5).value = 'Project_name'
		
		# Add CXCR3 for column 6
		dispatch_sheet.cell(row=row, column=6).value = 'Company_name'
		
		# Add XSel15min for column 12
		dispatch_sheet.cell(row=row, column=12).value = 'XSel15min'
		
		# Add 'no_salt' for column 14
		dispatch_sheet.cell(row=row, column=14).value = 'no_salt'		
		
		# Get cells only for NBID column
		NBID_value = [dispatch_sheet.cell(row=row, column=col).value for col in range(1, 40) if col != None and col == 7]	
		
		#print "NBID value is %r" % NBID_value
		
		
		#If cell is empty then store row
		if NBID_value == [None]:
									
			#print "Empty row number is: %d" % row
			
			# Now place from word doc into Excel 
			for key, value in table_row.iteritems():
				
				#print "They key is: %s" % key
				#print "The value is: %s" % value
				
				#print "row number is: %d" % row
				
				if key == 'Lab ID':
					dispatch_sheet.cell(row=row, column=7).value = value
				
				if key == 'Purity\n1dp':
					dispatch_sheet.cell(row=row, column=10).value = value
				
				if 'Rt' in key:
					dispatch_sheet.cell(row=row, column=11).value = value
					
				if 'Biosub' in key:
					dispatch_sheet.cell(row=row, column=2).value = value								
					
				if 'Store amount' in key:
					dispatch_sheet.cell(row=row, column=4).value = value
					
					if value == '':
						dispatch_sheet.cell(row=row, column=3).value = 'N'
						dispatch_sheet.cell(row=row, column=4).value = '-'
					
					else:
						dispatch_sheet.cell(row=row, column=3).value = 'Y'
					
				if 'DupOf' in key:
					dispatch_sheet.cell(row=row, column=19).value = value
					
				if 'Stereo' in key:
					dispatch_sheet.cell(row=row, column=21).value = value
					
				if 'ion' in key:
					dispatch_sheet.cell(row=row, column=13).value = value
					
				if 'Comments' in key:
					dispatch_sheet.cell(row=row, column=20).value = value
					
				if 'even' in key:
					dispatch_sheet.cell(row=row, column=29).value = value
				
				if 'odd' in key:
					dispatch_sheet.cell(row=row, column=30).value = value	
							
			
			break	

def create_dictionary(sheet, dispatch_sheet, date, day):
	# Headers start from row 7. Seems that sheet.max_row represents last row with something in it
	for row in range(7, sheet.max_row):	

		#print "row is: %s" % row	

		# Count filled cells
		filled = 0		

		# List comprehension		
		single_row = [sheet.cell(row=row, column=col).value for col in range (1,16)]

		#print "\n" + str(single_row) + "\n"


		# Store headers as keys
		if row == 7:
			keys = single_row

		else:
			# Check how many empty cells are in a row
			for j in single_row:
				if j != None:
					filled += 1
			#print "\nFilled cells in this row are: %s\n" % filled

			# If less than 1 filled cell, discount row
			if filled < 1:
				single_row = []


			# Only zip row with keys list if they are the same length. Avoids mis-matched key-value pairs
			if len(single_row) == len(keys):				
				row_key_value_pairs = dict(zip(keys, single_row))				
				#print "row %d is: %s " % (row, row_key_value_pairs)			
				
				if day == "Wednesday":
					# Now pass row to dispatch sheet
					wednesday_dump(row_key_value_pairs, dispatch_sheet)			


		if sheet.cell(row=row, column=col).value == None:
			continue

		#print "\n" + str(keys) + "\n"



# Function to grab a row at a time from each person's dispatch sheet
def analyse_dispatch_sheets(Macro_docs, dispatch_sheet, date, Day, day_number):
	
	# Variables to keep track of how many and which files are still open
	open_file_count = 0
	open_docs = []	
	
	# This bit of code (move to separate function) analyses the word_docs list for any copies... ie files created if the orginal file is open. As found that even if move the temp file, still can not manipulate the file that is open. So have to create a copy and manipulate the copy instead.
	# Manipulating the copy does work! (Test if still need to delete the temp as well as create copy - No, do not have to move the temp. Can just exclude from the word_list array)
	
	# Exclude temp files
	Macro_docs = [doc for doc in Macro_docs if not doc.startswith('~')]
	
	# Use copy of files that are already open. Delete after finished with them
	for doc in Macro_docs:		
				
		if 'copy' in doc:			
			# Get name of file before copy
			name = doc.rsplit('_copy', 1)[0]			
			# Now ignore/remove that file from word_docs list
			name = name + '.xlsm'
			#print name			
			Macro_docs.remove(name)
	
	
	for doc in Macro_docs:		
		
		print "doc is: %s" % doc
		
		# Open individual dispatch sheet. Opening macro xlsm, so to avoid corrupting file need to send other arguments? : https://stackoverflow.com/questions/17675780/how-to-save-xlsm-file-with-macro-using-openpyxl
		wb = openpyxl.load_workbook(filename=doc, read_only=False, keep_vba=True)
		sheet = wb.active 
					
		#print "Individual sheet is currently: %s" % sheet
		#print "Master dispatch sheet is: %s" % dispatch_sheet
		
		create_dictionary(sheet, dispatch_sheet, date, Day)
		
		
		# save document 
		try:
			wb.save(doc)
			# If doc contains 'copy', delete it
			if 'copy' in doc:
				os.remove("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % doc)
		
		except:
			
			print "\nSorry, the file %s is still open..." % doc
			
			# Update local open_file_count
			open_file_count += 1
			# Push to open_doc array
			open_docs.append(doc)	
		
		# If more than one document is open, offer user choices as to copy files or not. Then quit the program
		if open_file_count > 0:
				
			naughty_list = ' and '.join(open_docs)
		
			if now_time >= time (10,45):
		
				print"\nPlease check with %s to see if they are finished with their word document(s)...\n" % naughty_list
			
				return_value = raw_input("Please type 'yes' if they are finished, 'no' if the individual(s) requires more time: ")
			
				if return_value == 'yes':
					check = raw_input("Would you like to create a copy of the file(s)?: ")					
					if check == 'yes':
						# Try to copy and rename open files (and then delete copy afte finished with it?) and run again. Would be great here to be able to automatically run script again
						for file in open_docs:
							copy = file.replace('.xlsm','') + '_copy.xlsm'								
							copyfile("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % file, "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2/%s" % copy)
							print "\nCopy %s has been made.\n" % copy	
						print "\nPlease re-run the script\n"					
					else:
						print "\nOk, please try and run the script again in a few moments\n"
			
				else:
					print "\nOk, please try again later\n"
		
			else:
				print "\nPlease allow more time for completion of dispatch sheets\n"		
		
			quit()	
		
		

# Empty array for storing dictionary
keys =[]	
		
# Get today's date
print('Date today: %s' % datetime.date.today())
date = datetime.date.today()

# Determines and prints actual day of the week and month. See here for documentation: https://docs.python.org/3/library/datetime.html#strftime-strptime-behavior
now = datetime.datetime.now()
now_time = now.time()

# Get the day of the week
Day = now.strftime("%A")
#print "The day of the week is: %s" % Day

# Get the month of the year
Month = now.strftime("%B")
#print"The month is: %s" % Month

# Get the current year
Year = now.strftime("%Y")
#print"The year is: %s" % Year

# What is the day of the month
day_number = now.strftime("%d")
#print "The day of the month is: %s" % day_number

# Gets all Excel macro files in specified directory
Macro_docs = [each for each in os.listdir("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version2") if each.endswith('.xlsm')]

# If no dispatch sheets available, quit program
if not Macro_docs:             
	print "\nNo Dispatch sheets available for analysis\n"	
	quit()

print "\nCurrent dispatch sheets are: %s\n" % Macro_docs
	
#Day = "Thursday"
Day = raw_input("\nPlease provide the day of the week: ")

if Day.isalpha() == False:
	print "\nPlease provide a valid day of the week\n"


# Wednesday
if Day == 'Wednesday':
	
	#print "It is Wednesday"	
	
	# Open Master Dispatch sheet
	wb_dispatch = openpyxl.load_workbook(filename='control_files/DISPATCH_TEMPLATE_new.xlsm', read_only=False, keep_vba=True)
	dispatch_sheet = wb_dispatch["ChemOffice1"]
	
	# If Wednesday then need to create new dispatch excel sheet based on timedate stamp
	#print "Saving new spreadsheet in appropiate folder... "
	
	# First create appropiate directory if does not exist...
	if not os.path.exists("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version3/%s/%s_%s" % (Year, Month, Year)):
		os.makedirs("/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version3/%s/%s_%s" % (Year, Month, Year))	
	
			
	# Add 1 to day number (dispatch day is next day ie Wednesday)
	#day_number = int(day_number) + 1
	#day_number = "%02d" % day_number
	
	analyse_dispatch_sheets(Macro_docs, dispatch_sheet, date, Day, day_number)	
	
	#Add source_sheet names to separate sheet of dispatch template
	copy_sheet_ids = wb_dispatch["Dispatch_Sheets"]	
	copy_across_sheet_ids(copy_sheet_ids, Macro_docs)	 	
	
	#print "day_number is: %s" % day_number	
	
	# Note that day number needs to be +1
	#save_location = "s:/Receptos/Dispatch/%s/%s_%s/%s_%s_%s.xlsx" % (Year, Month, Year, day_number, Month, Year) 
	save_location = "/Users/matthewcrittall/Documents/Programming/Webdev/Python/Projects/Dispatch_Python/dispatch_client/Version3/%s/%s_%s/%s_%s_%s.xlsm" % (Year, Month, Year, day_number, Month, Year)		
	wb_dispatch.save(save_location)
	#copy_in_structures(Macro_docs, save_location, Day) #Perhaps only run this function at end of day i.e. before uplaoding spreadsheet. Seems like could be a bit buggy... so don't want to screw up in morning ie keep seeing 'runtime error 13 type mismatch'. Perhaps because old file has not closed? Trying to open ~file?
	#Perhaps opening and closing comm link for every Doc is also causing problems. Perhaps better to pass Excel array of files and let VBA iterate through the array rather than Python
		 	
	
	print"\nProcess completed successfully\n"
	
	#Finish
	quit()








	
	
	
	
	
		
		